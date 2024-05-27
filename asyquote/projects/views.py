import json

from django.db.models import Max
import re
from django.shortcuts import render, redirect, get_object_or_404
from .forms import ProjectForm
from .models import Project, SectionQuote, ServicesQuote, PricesQuote, Notes
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import render_to_string
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from wagtail.blocks import StreamValue
from ..clients.models import Client
from ..products.models import Products
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, numbers


@login_required
def project_list(request):
    state = "EM ESPERA"
    projects = Project.objects.filter(user=request.user, state=state)
    form = ProjectForm(user=request.user)
    form.fields['client'].queryset = form.fields['client'].queryset.order_by('name')
    return render(request, 'builder/builder_projects.html', {'projects': projects, 'form': form})


def edit_project(request, key):
    project = get_object_or_404(Project, key=key)
    if project.user != request.user:
        return render(request, 'builder/project_not_found.html')
    clients = Client.objects.filter(user=request.user)
    sections = SectionQuote.objects.filter(project_key=key).order_by('id')
    services = ServicesQuote.objects.filter(project_key=key).order_by('id')
    prices = PricesQuote.objects.filter(project_key=key).order_by('id')
    notes = Notes.objects.filter(project_key=key)

    return render(request, 'builder/edit_project.html',
                  {'project': project, 'clients': clients, 'sections': sections, 'services': services, 'prices': prices,
                   'notes': notes})


def create_project(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES)
        if form.is_valid():
            last_project = Project.objects.order_by('-id').first()
            if last_project:
                last_project_id = last_project.id
            else:
                last_project_id = 0

            new_id = last_project_id + 1
            form.cleaned_data['id'] = new_id

            project = form.save(commit=False)
            project.user = request.user  # Set the user
            project.state = 'EM ESPERA'
            project.save()
            key = project.key
            max_section_count = SectionQuote.objects.filter(project_key=key).aggregate(Max('section_count'))
            next_section_count = max_section_count['section_count__max'] + 1 if max_section_count[
                                                                                    'section_count__max'] is not None else 1
            SectionQuote.objects.create(project_key=key, section_count=next_section_count)

            # services_creation
            ServicesQuote.objects.create(project_key=key, service_count=1, section_key=next_section_count)

            # prices_creation
            PricesQuote.objects.create(project_key=key, prices_count=1, services_key=1, section_key=next_section_count)

            messages.success(request, 'Projeto criado com sucesso!')
            return redirect('project_list')
        else:

            messages.error(request, 'Erro ao criar projeto. Verifique se as informações estão corretas.')
            return redirect('project_list')
    else:
        form = ProjectForm()

    return render(request, 'builder/builder_projects.html', {'form': form})


def filter_projects(request):
    state = request.GET.get('state', None)
    projects = Project.objects.filter(user=request.user)

    if state:
        projects = projects.filter(state=state, user=request.user)

    html = render_to_string('builder/project_list_partial.html', {'projects': projects})
    return HttpResponse(html)


def filter_projects_data(request):
    state = request.GET.get('state', None)
    projects = Project.objects.filter(user=request.user)

    if state:
        projects = projects.filter(state=state, user=request.user)
    html = render_to_string('builder/project_list_data_partial.html',
                            {'projects': projects})
    return HttpResponse(html)


def list_projects_table(request):
    state = "EM EXECUÇÃO"
    projects = Project.objects.filter(user=request.user, state=state)
    return render(request, 'builder/project_list.html', {'projects': projects})


def download_excel(request):
    state = request.GET.get('state')
    if state == None:
        state = "EM EXECUÇÃO"
        projects = Project.objects.filter(state=state, user=request.user).order_by('quote_number')
    elif state:
        projects = Project.objects.filter(state=state, user=request.user).order_by('quote_number')
    else:
        projects = Project.objects.filter(user=request.user).order_by('quote_number')
    if not projects:
        messages.info(request, "Não foi encontrado nehum projeto com estas condições")
        return redirect('list_projects_table')
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add heading row
    heading_row = ['Ref. Orçamento', 'Título', 'Cliente', 'Localização', 'Custo', 'Cobrado', 'Percentagem', 'Estado']
    ws.append(heading_row)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # Add data rows
    for project in projects:
        ws.append([project.quote_number, project.title, project.client.name, project.address,
                   f"{project.total_cost():.2f} €",
                   f"{project.total_charged():.2f} €", f"{project.profit_percentage():.2f} %", project.state])

    total_cost = sum(project.total_cost() for project in projects)
    total_charged = sum(project.total_charged() for project in projects)
    if total_cost != 0:
        total_profit = ((total_charged - total_cost) / total_cost) * 100
    else:
        total_profit = 0
    total_row = ['Total', '', '', '', f"{total_cost:.2f} €", f"{total_charged:.2f} €", f"{total_profit:.2f} %"]
    ws.append(total_row)
    total_cell = ws.cell(row=len(projects) + 2, column=1)
    total_cell.fill = PatternFill(start_color='05af50', end_color='05af50', fill_type='solid')
    total_cell.font = Font(color='FFFFFF', bold=True)

    for col in range(1, len(heading_row) + 1):
        max_length = 0
        for row in range(1, len(projects) + 3):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Save the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if state:
        filename = f"projetos_{state.lower().replace(' ', '_')}.xlsx"
    else:
        filename = "projetos.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def delete_project(request, key):
    project = get_object_or_404(Project, key=key)
    sections = SectionQuote.objects.filter(project_key=key)
    services = ServicesQuote.objects.filter(project_key=key)
    prices = PricesQuote.objects.filter(project_key=key)
    notes = Notes.objects.filter(project_key=key)

    if request.method == 'POST':
        project.delete()
        prices.delete()
        services.delete()
        sections.delete()
        notes.delete()
        return redirect('project_list')
    return render(request, 'builder/delete_project.html', {'project': project})


def update_project(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    if request.method == 'POST':

        updated_title = request.POST.get('update_title')
        updated_address = request.POST.get('update_address')
        updated_state = request.POST.get('update_state')
        updated_client_id = request.POST.get('update_project_client')
        if updated_title and updated_address:
            project.title = updated_title
            project.address = updated_address
            updated_client = get_object_or_404(Client, pk=updated_client_id)
            project.state = updated_state
            project.client = updated_client
            project.save()
            return JsonResponse({'success': 'Informações atualizadas com sucesso'})
    else:
        return JsonResponse({'error': 'Informação errada'})


def delete_fields_quote(request):
    if request.method == 'GET':
        action = request.GET.get('action')
        key = request.GET.get('key')
        section_count = request.GET.get('section_count')
        service_count = request.GET.get('service_count')
        price_count = request.GET.get('price_count')
        try:
            if action == 'drop-section':
                num_sections = SectionQuote.objects.filter(project_key=key, visible=True).count()
                if num_sections > 1:
                    section = SectionQuote.objects.get(project_key=key, section_count=section_count)
                    services = ServicesQuote.objects.filter(project_key=key, section_key=section_count)
                    prices = PricesQuote.objects.filter(project_key=key, section_key=section_count)
                    section.visible = False
                    for service in services:
                        service.visible = False

                    for price in prices:
                        price.visible = False
                    section.save()
                    return JsonResponse({'status': 'success', 'message': 'Section hidden successfully.'})
                else:
                    return JsonResponse({'status': 'error', 'message': 'Cannot delete the only section.'})

            elif action == 'drop-service':
                num_services = ServicesQuote.objects.filter(project_key=key, visible=True,
                                                            section_key=section_count).count()
                print("num_services:", num_services)
                if num_services > 1:
                    service = ServicesQuote.objects.get(project_key=key, section_key=section_count,
                                                        service_count=service_count, visible=True)
                    print(service)
                    service.visible = False
                    service.save()
                    prices = PricesQuote.objects.filter(project_key=key, section_key=section_count,
                                                        services_key=service_count, visible=True)

                    for price in prices:
                        print("price_name:", price.name)
                        price.visible = False
                        price.save()

                    return JsonResponse({'status': 'success', 'message': 'Service hidden successfully.'})
                else:
                    return JsonResponse({'status': 'error', 'message': 'Cannot delete the only service.'})

            elif action == 'drop-price':
                num_prices = PricesQuote.objects.filter(project_key=key, visible=True, section_key=section_count,
                                                        services_key=service_count).count()
                if num_prices > 1:
                    price = PricesQuote.objects.get(project_key=key, section_key=section_count,
                                                    services_key=service_count,
                                                    prices_count=price_count)
                    price.visible = False
                    price.save()
                    return JsonResponse({'status': 'success', 'message': 'Price hidden successfully.'})
                else:
                    return JsonResponse({'status': 'error', 'message': 'Cannot delete the only price.'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid action.'})

        except SectionQuote.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Section not found.'})
        except ServicesQuote.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Service not found.'})
        except PricesQuote.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Price not found.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


def save_quote_url(request):
    if request.method == 'GET':
        action = request.GET.get('action')
        key = request.GET.get('key')
        project = get_object_or_404(Project, key=key)
        clients = Client.objects.all()
        sections = SectionQuote.objects.filter(project_key=key).order_by('id')
        services = ServicesQuote.objects.filter(project_key=key).order_by('id')
        prices = PricesQuote.objects.filter(project_key=key).order_by('id')
        notes = Notes.objects.filter(project_key=key)
        if action == 'add-section':
            # section_creation
            max_section_count = SectionQuote.objects.filter(project_key=key).aggregate(Max('section_count'))
            next_section_count = max_section_count['section_count__max'] + 1 if max_section_count[
                                                                                    'section_count__max'] is not None else 1
            SectionQuote.objects.create(project_key=key, section_count=next_section_count)

            # services_creation
            ServicesQuote.objects.create(project_key=key, service_count=1, section_key=next_section_count)

            # prices_creation
            PricesQuote.objects.create(project_key=key, prices_count=1, services_key=1, section_key=next_section_count)

        if action == 'add-service':
            # get the latest section count
            max_section_count = SectionQuote.objects.filter(project_key=key).aggregate(Max('section_count'))
            next_section_count = max_section_count['section_count__max']

            # get the latest service count for the current section
            max_services_count = ServicesQuote.objects.filter(project_key=key,
                                                              section_key=next_section_count).aggregate(
                Max('service_count'))
            next_service_count = max_services_count['service_count__max'] + 1 if max_services_count[
                                                                                     'service_count__max'] is not None else 1

            # services_creation
            ServicesQuote.objects.create(project_key=key, service_count=next_service_count,
                                         section_key=next_section_count)

            # get the latest price count for the current service
            max_prices_count = PricesQuote.objects.filter(project_key=key, services_key=next_service_count,
                                                          section_key=next_section_count).aggregate(
                Max('prices_count'))
            next_prices_count = max_prices_count['prices_count__max'] + 1 if max_prices_count[
                                                                                 'prices_count__max'] is not None else 1

            # prices_creation
            PricesQuote.objects.create(project_key=key, prices_count=next_prices_count, services_key=next_service_count,
                                       section_key=next_section_count)

        if action == 'add-price':
            # get the latest section count
            max_section_count = SectionQuote.objects.filter(project_key=key).aggregate(Max('section_count'))
            next_section_count = max_section_count['section_count__max']

            # get the latest service count for the current section
            max_services_count = ServicesQuote.objects.filter(project_key=key,
                                                              section_key=next_section_count).aggregate(
                Max('service_count'))
            next_service_count = max_services_count['service_count__max']

            # get the latest price count for the current service
            max_prices_count = PricesQuote.objects.filter(project_key=key, services_key=next_service_count,
                                                          section_key=next_section_count).aggregate(
                Max('prices_count'))
            next_prices_count = max_prices_count['prices_count__max'] + 1 if max_prices_count[
                                                                                 'prices_count__max'] is not None else 1

            # prices_creation
            PricesQuote.objects.create(project_key=key, prices_count=next_prices_count, services_key=next_service_count,
                                       section_key=next_section_count)

        form_html = render_to_string('builder/edit_project_partial.html',
                                     {'project': project, 'clients': clients, 'sections': sections,
                                      'services': services, 'prices': prices,
                                      'notes': notes})

        return JsonResponse({'form_html': form_html})


def save_quote_data(request):
    if request.method == 'GET':
        action = request.GET.get('action')
        key = request.GET.get('key')
        value = request.GET.get('value')
        section_key = request.GET.get('section_count')
        service_key = request.GET.get('service_count')
        price_key = request.GET.get('price_count')

        clients = Client.objects.all()
        project = get_object_or_404(Project, key=key)
        sections = SectionQuote.objects.filter(project_key=key).order_by('id')
        services = ServicesQuote.objects.filter(project_key=key).order_by('id')
        prices = PricesQuote.objects.filter(project_key=key).order_by('id')
        notes = Notes.objects.filter(project_key=key)
        if action == 'sections':
            section_changed = sections.get(section_count=section_key, visible=True)
            section_changed.name = value
            section_changed.save()
            sections = SectionQuote.objects.filter(project_key=key, visible=True).order_by('id')

        elif action == 'services':
            services_changed = services.get(section_key=section_key, service_count=service_key, visible=True)
            services_changed.name = value
            services_changed.save()
            services = ServicesQuote.objects.filter(project_key=key, visible=True).order_by('id')

        elif action == 'quantities':
            services_changed = services.get(section_key=section_key, service_count=service_key, visible=True)
            services_changed.quantity = value
            services_changed.save()
            services = ServicesQuote.objects.filter(project_key=key, visible=True).order_by('id')

        elif action == 'description':
            prices_changed = prices.get(section_key=section_key, services_key=service_key, prices_count=price_key,
                                        visible=True)
            prices_changed.description = value
            prices_changed.save()
            prices = PricesQuote.objects.filter(project_key=key, visible=True).order_by('id')
        elif action == 'cost':
            prices_changed = prices.get(section_key=section_key, services_key=service_key, prices_count=price_key,
                                        visible=True)
            prices_changed.cost = value
            prices_changed.save()
            prices = PricesQuote.objects.filter(project_key=key, visible=True).order_by('id')
        elif action == 'charged':
            prices_changed = prices.get(section_key=section_key, services_key=service_key, prices_count=price_key,
                                        visible=True)
            prices_changed.charged = value
            prices_changed.save()
            prices = PricesQuote.objects.filter(project_key=key, visible=True).order_by('id')
        elif action == 'notes':
            notes_changed = notes.get(project_key=key)
            notes_changed.notes = value
            notes_changed.save()
            notes = Notes.objects.filter(project_key=key)

        form_html = render_to_string('builder/edit_project_partial.html',
                                     {'sections': sections,
                                      'services': services, 'prices': prices,
                                      'notes': notes})

        return JsonResponse({'form_html': form_html})


def filter_edit_products(request):
    if request.method == 'GET':

        key = request.GET.get('key')
        value = request.GET.get('value')

        products = Products.objects.all()
        sections = SectionQuote.objects.filter(project_key=key).order_by('id')
        services = ServicesQuote.objects.filter(project_key=key).order_by('id')
        prices = PricesQuote.objects.filter(project_key=key).order_by('id')
        notes = Notes.objects.filter(project_key=key)

        if value != '':
            filtered_products = products.filter(title__icontains=value).order_by('-id')
        else:
            filtered_products = None

        form_html = render_to_string('builder/edit_project_products.html',
                                     {'sections': sections,
                                      'services': services, 'prices': prices,
                                      'notes': notes, 'filtered_products': filtered_products})

        return JsonResponse({'form_html': form_html})


def download_project_quote(request, project_key):
    project = Project.objects.get(key=project_key, user=request.user)
    notes = Notes.objects.get(project_key=project_key)
    # projects = Project.objects.filter(user=request.user).order_by('quote_number')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 6.0
    ws.column_dimensions['B'].width = 41.83
    ws.column_dimensions['C'].width = 6.0
    ws.column_dimensions['D'].width = 4.0
    ws.column_dimensions['E'].width = 9.17
    ws.column_dimensions['F'].width = 10.17
    ws.print_title_rows = '1:4'
    ws.print_title_cols = 'A:F'

    # sub_heading
    sub_heading_row = ['Proposta']
    ws.append(sub_heading_row)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=6)
    ws.cell(row=1, column=1).value = 'Proposta'
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='right', vertical='center')

    # project info
    project_address = project.address
    project_number = project.quote_number
    project_info = ['Obra', project_address, '', '', 'Proposta Nº', project_number]
    for col, value in enumerate(project_info, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = value
        cell.font = Font(size=8, bold=True) if value in ['Obra', 'Proposta Nº'] else Font(size=8)
        cell.border = Border(top=Side(border_style="thin"), right=Side(border_style="dotted"),
                             bottom=Side(border_style="dotted"))
        if value in ['Obra', 'Proposta Nº']:
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            if value in 'Obra':
                cell.border = Border(left=Side(border_style="thin"), top=Side(border_style="thin"),
                                     bottom=Side(border_style="dotted"), right=Side(border_style="dotted"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        elif value == project_number:
            cell.border = Border(top=Side(border_style="thin"), right=Side(border_style="thin"),
                                 bottom=Side(border_style="dotted"))
            cell.alignment = Alignment(horizontal="center", vertical="center")

    project_client = project.client.name
    current_date = datetime.now()
    formatted_date = current_date.strftime("%d/%m/%Y")
    project_info_2 = ['Cliente', project_client, '', '', 'Data', formatted_date]
    for col, value in enumerate(project_info_2, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = value
        cell.font = Font(size=8, bold=True) if value in ['Cliente', 'Data'] else Font(size=8)
        cell.border = Border(right=Side(border_style="dotted"), bottom=Side(border_style="thin"))
        if value in ['Cliente', 'Data']:
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            if value in 'Cliente':
                cell.border = Border(left=Side(border_style="thin"), bottom=Side(border_style="thin"),
                                     right=Side(border_style="dotted"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        elif value == formatted_date:
            cell.border = Border(bottom=Side(border_style="thin"), right=Side(border_style="thin"))
            cell.alignment = Alignment(horizontal="center", vertical="center")

    start_column = 3
    end_column = 4
    start_row = 3
    end_row = 4
    start_cell = get_column_letter(start_column) + str(start_row)
    end_cell = get_column_letter(end_column) + str(end_row)
    ws.merge_cells(start_cell + ":" + end_cell)
    merged_cell_range = start_cell + ":" + end_cell
    for merged_cell in ws[merged_cell_range]:
        for cell in merged_cell:
            cell.border = Border(bottom=Side(border_style="thin"), top=Side(border_style="thin"),
                                 right=Side(border_style="dotted"))

    # separator
    separator = ['', '', '', '', '', '']
    for col, value in enumerate(separator, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = value

    # heading
    heading_row = ['Artigo', 'Descrição do Artigo', 'Quant.', 'Un.', 'Valor']
    ws.append(heading_row)
    ws.merge_cells('A6:A7')
    ws.merge_cells('B6:B7')
    ws.merge_cells('C6:C7')
    ws.merge_cells('D6:D7')
    ws.merge_cells('E6:F6')

    ws.cell(row=7, column=5).value = "Unitário"
    ws.cell(row=7, column=6).value = "Parcial"

    alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = alignment
            cell.font = Font(size=8, bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            if cell.column == 1:
                cell.border = Border(left=Side(border_style="thin"), top=Side(border_style="thin"),
                                     right=Side(border_style="dotted"), bottom=Side(border_style="dotted"))
            elif cell.column == 6 and cell.row != 7:
                cell.border = Border(right=Side(border_style="thin"), top=Side(border_style="thin"),
                                     bottom=Side(border_style="dotted"))
            elif cell.row == 7 and cell.column == 6:
                cell.border = Border(right=Side(border_style="thin"), bottom=Side(border_style="dotted"))
            elif cell.row == 7 and cell.column == 5:
                cell.border = Border(top=Side(border_style="dotted"), bottom=Side(border_style="dotted"),
                                     right=Side(border_style="dotted"))
            else:
                cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="dotted"),
                                     right=Side(border_style="dotted"))

    current_row = 9
    count = 1
    count_2 = 0
    total_quote = 0
    green_fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")
    sections = SectionQuote.objects.filter(project_key=project_key).order_by('section_count')
    for section in sections:
        sum_columns = 0
        sum_prices = 0
        count_2 = 0
        if section.name:
            article_section = ws.cell(row=current_row, column=1, value=count)
            article_section.font = Font(size=10, bold=True)
            article_section.alignment = Alignment(horizontal="right", vertical="center")
            section_description = ws.cell(row=current_row, column=2, value=section.name)
            section_description.alignment = Alignment(wrap_text=True, vertical="center")
            section_description.font = Font(size=10, bold=True)
            services = ServicesQuote.objects.filter(project_key=project_key,
                                                    section_key=section.section_count).order_by(
                'service_count')
            current_row += 1
            for service in services:
                if service.name:
                    count_2 += 1
                    article_string = str(count) + "." + str(count_2)
                    article_number = float(article_string)
                    service_description = ws.cell(row=current_row, column=2, value=service.name)
                    service_description.alignment = Alignment(wrap_text=True, vertical="center")
                    service_description.font = Font(size=10)
                    article_cell = ws.cell(row=current_row, column=1, value=article_number)
                    article_cell.alignment = Alignment(horizontal="right", vertical="center")
                    article_cell.font = Font(size=10)
                    quantity_string = service.quantity
                    quantity_number = ''
                    quantity_measure = 'un'
                    quantity_letter = ''
                    quantity_flag = 0
                    if quantity_string:
                        for i in quantity_string:
                            if not (i.isdigit()):
                                quantity_flag = 1
                            if quantity_flag == 1:
                                quantity_letter += i.strip()
                            else:
                                quantity_number += i
                        if quantity_letter != '':
                            quantity_measure = quantity_letter
                        if quantity_number and int(quantity_number) > 0:
                            quantity_number = int(quantity_number)
                        else:
                            quantity_number = 1
                        quantity_number_cell = ws.cell(row=current_row, column=3, value=quantity_number)
                        quantity_number_cell.number_format = "0.00"
                        quantity_number_cell.alignment = Alignment(horizontal="right", vertical="center")
                        quantity_number_cell.font = Font(size=8)
                        quantity_measure_cell = ws.cell(row=current_row, column=4, value=quantity_measure)
                        quantity_measure_cell.alignment = Alignment(horizontal="center", vertical="center")
                        quantity_measure_cell.font = Font(size=8)
                    prices = PricesQuote.objects.filter(project_key=project_key, section_key=section.section_count,
                                                        services_key=service.service_count).order_by('prices_count')
                    for price in prices:
                        if price.charged and quantity_number:
                            sum_columns += quantity_number * int(price.charged)
                        if price.charged:
                            sum_prices += int(price.charged)

                    sum_price_cell = ws.cell(row=current_row, column=5, value=sum_prices)
                    sum_price_cell.number_format = '€#,##0.00'
                    sum_price_cell.font = Font(size=8)
                    sum_price_cell.alignment = Alignment(horizontal="right", vertical="center")
                    formula_cell = ws.cell(row=current_row, column=6, value=f"=C{current_row}*E{current_row}")
                    formula_cell.number_format = '€#,##0.00'
                    formula_cell.font = Font(size=8)
                    formula_cell.alignment = Alignment(horizontal="center", vertical="center")

                    current_row += 1
                    sum_prices = 0

            total_quote += sum_columns
            current_row += 1
            total_title = "Total Capítulo " + str(count)
            total_chapter = ws.cell(row=current_row, column=6, value=sum_columns)
            total_chapter.number_format = '€#,##0.00'
            total_chapter.font = Font(size=8)
            total_chapter_title = ws.cell(row=current_row, column=2, value=total_title)
            total_chapter_title.font = Font(size=10)
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = green_fill

            current_row += 2
            count += 1

    total_quote_text = ws.cell(row=current_row, column=2, value="Total da Proposta")
    total_quote_number = ws.cell(row=current_row, column=6, value=total_quote)
    total_quote_text.font = Font(size=8, bold=True)
    total_quote_number.font = Font(size=8, bold=True)
    total_quote_number.number_format = '€#,##0.00'
    total_quote_number.alignment = Alignment(horizontal="center")
    for col in range(1, 7):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if col == 1:
            cell.border = Border(left=Side(border_style="thin"), bottom=Side(border_style="thin"),
                                 right=Side(border_style="dotted"))
        elif col == 6:
            cell.border = Border(right=Side(border_style="thin"), bottom=Side(border_style="thin"),
                                 left=Side(border_style="dotted"))
        else:
            cell.border = Border(bottom=Side(border_style="thin"), right=Side(border_style="dotted"))

    for row in range(8, current_row):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.border = Border(left=Side(border_style="thin"), bottom=Side(border_style="dotted"),
                                     right=Side(border_style="dotted"))
            elif col == 6:
                cell.border = Border(bottom=Side(border_style="dotted"), right=Side(border_style="thin"))
            else:
                cell.border = Border(bottom=Side(border_style="dotted"), right=Side(border_style="dotted"))

    if notes.notes:
        notas = ws.cell(row=current_row + 1, column=1, value=notes.notes)
        notas.font = Font(size=10)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "projeto_" + project_number + ".xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def create_fields_quote(request):
    if request.method == 'GET':

        action = request.GET.get('action')
        key = request.GET.get('key')
        section_key = request.GET.get('section_count')
        service_key = request.GET.get('service_count')
        price_key = request.GET.get('price_count')
        next_id = request.GET.get('next_id')

        if action == 'add-service':
            next_services = ServicesQuote.objects.filter(project_key=key, section_key=section_key,
                                                         service_count=next_id, visible=True)
            max_prices_count = \
                PricesQuote.objects.filter(project_key=key, section_key=section_key, services_key=next_id) \
                    .aggregate(Max('prices_count'))['prices_count__max']
            next_prices_count = (max_prices_count or 0) + 1
            for service in next_services:
                print(service.name)
            if next_services:
                print("ola")
            else:
                ServicesQuote.objects.create(project_key=key, service_count=next_id,
                                             section_key=section_key)
                PricesQuote.objects.create(project_key=key, prices_count=next_prices_count,
                                           services_key=next_id,
                                           section_key=section_key)

        sections = SectionQuote.objects.filter(project_key=key).order_by('id')
        services = ServicesQuote.objects.filter(project_key=key).order_by('id')
        prices = PricesQuote.objects.filter(project_key=key).order_by('id')
        notes = Notes.objects.filter(project_key=key)
        form_html = render_to_string('builder/edit_project_partial.html',
                                     {'sections': sections,
                                      'services': services, 'prices': prices,
                                      'notes': notes})

        return JsonResponse({'form_html': form_html})
