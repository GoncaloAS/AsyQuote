"""The Excel exports: what the client actually receives."""
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from .factories import add_line, add_service, make_user, simple_quote

pytestmark = pytest.mark.django_db

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def workbook_from(response):
    return load_workbook(BytesIO(response.content))


class TestQuoteExport:
    def download(self, client, project):
        return client.get(reverse("download_project_quote", kwargs={"project_key": project.key}))

    def test_returns_a_spreadsheet_named_after_the_quote(self, client):
        user = make_user()
        project, _, _ = simple_quote(user, quote_number="07")
        client.force_login(user)

        response = self.download(client, project)

        assert response.status_code == 200
        assert response["Content-Type"] == XLSX
        assert "projeto_07.xlsx" in response["Content-Disposition"]

    def test_carries_the_client_and_the_site(self, client):
        user = make_user()
        project, _, _ = simple_quote(user)
        client.force_login(user)

        sheet = workbook_from(self.download(client, project)).active

        assert sheet["B3"].value == project.address
        assert sheet["B4"].value == project.client.name
        assert sheet["F3"].value == project.quote_number

    def test_header_repeats_on_every_printed_page(self, client):
        """Without this a multi-page quote loses its header when printed."""
        user = make_user()
        project, _, _ = simple_quote(user)
        client.force_login(user)

        sheet = workbook_from(self.download(client, project)).active

        assert sheet.print_title_rows in ("1:4", "$1:$4")

    def test_line_totals_are_formulas_the_client_can_edit(self, client):
        user = make_user()
        project, _, _ = simple_quote(user)
        client.force_login(user)

        sheet = workbook_from(self.download(client, project)).active
        cells = [c.value for col in sheet.iter_cols(min_col=6, max_col=6) for c in col]
        formulas = [v for v in cells if str(v).startswith("=")]

        assert formulas, "expected at least one =C*E line total"
        assert all(f.startswith("=C") and "*E" in f for f in formulas)

    def test_articles_are_numbered_by_position_in_the_tree(self, client):
        user = make_user()
        project, chapter, service = simple_quote(user)
        add_service(project, chapter, "Segundo serviço", position=2)
        client.force_login(user)

        sheet = workbook_from(self.download(client, project)).active
        cells = [c.value for col in sheet.iter_cols(min_col=1, max_col=1) for c in col]
        articles = [v for v in cells if isinstance(v, (int, float))]

        assert 1 in articles, "the chapter should be article 1"
        assert 1.1 in articles, "its first service should be 1.1"
        assert 1.2 in articles, "its second service should be 1.2"

    def test_chapter_subtotal_multiplies_quantity_by_price(self, client):
        user = make_user()
        project, _, _ = simple_quote(user)  # 10 m2, lines charged 150 + 250
        client.force_login(user)

        sheet = workbook_from(self.download(client, project)).active
        cells = [c.value for col in sheet.iter_cols(min_col=6, max_col=6) for c in col]
        values = [v for v in cells if isinstance(v, (int, float))]

        assert 4000 in values, "10 m2 x 400 charged should reach the chapter total"

    def test_a_hidden_line_is_left_out(self, client):
        user = make_user()
        project, chapter, service = simple_quote(user)
        extra = add_line(project, chapter, service, cost=10, charged=999, description="Apagada", position=3)
        extra.visible = False
        extra.save()
        client.force_login(user)

        sheet = workbook_from(self.download(client, project)).active
        text = " ".join(str(c.value) for row in sheet.iter_rows() for c in row)

        assert "Apagada" not in text

    def test_another_user_cannot_download_it(self, client):
        owner = make_user("owner")
        project, _, _ = simple_quote(owner)
        client.force_login(make_user("intruder"))

        assert self.download(client, project).status_code == 404


class TestProjectListExport:
    def test_lists_the_quotes_in_a_state_with_a_totals_row(self, client):
        user = make_user()
        project, _, _ = simple_quote(user, quote_number="11")
        project.state = "EM EXECUÇÃO"
        project.save()
        client.force_login(user)

        response = client.get(reverse("download_excel"), {"state": "EM EXECUÇÃO"})
        sheet = workbook_from(response).active
        rows = list(sheet.iter_rows(values_only=True))

        assert rows[0][0] == "Ref. Orçamento"
        assert any(r[0] == "11" for r in rows), "the quote should be listed"
        assert rows[-1][0] == "Total"

    def test_does_not_leak_another_user_quotes(self, client):
        other = make_user("other")
        theirs, _, _ = simple_quote(other, quote_number="99")
        theirs.state = "EM EXECUÇÃO"
        theirs.save()

        mine = make_user("mine")
        ours, _, _ = simple_quote(mine, quote_number="12")
        ours.state = "EM EXECUÇÃO"
        ours.save()
        client.force_login(mine)

        sheet = workbook_from(client.get(reverse("download_excel"), {"state": "EM EXECUÇÃO"})).active
        text = " ".join(str(c.value) for row in sheet.iter_rows() for c in row)

        assert "12" in text
        assert "99" not in text
