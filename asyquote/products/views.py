import os
import re
import time
from decimal import Decimal, DecimalException
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.views.generic import ListView, TemplateView
from openpyxl import load_workbook

from .forms import CategoryForm, ProductsForm, SupplierForm, UploadExcelForm
from .models import Category, Links, Products, Supplier


def superuser_required(view):
    """Catalogue writes are superuser-only, which was enforced only in templates."""
    return login_required(user_passes_test(lambda user: user.is_superuser)(view))


class NotFoundProductView(TemplateView):
    template_name = "404/404_products.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class ProductsListView(ListView):
    template_name = "products/products_page.html"
    model = Products
    paginate_by = 24
    context_object_name = "products"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_suppliers = Supplier.objects.all()
        all_categorys = Category.objects.all()
        form = ProductsForm()
        form_categories = CategoryForm()
        form_supplier = SupplierForm()
        form_excel = UploadExcelForm()
        context["all_suppliers"] = all_suppliers
        context["all_categorys"] = all_categorys
        context["form"] = form
        context["form_categories"] = form_categories
        context["form_supplier"] = form_supplier
        context["form_excel"] = form_excel
        return context

    def get_queryset(self):
        return Products.objects.all().order_by("-id")


@login_required
def filter_products(request):
    category_names = request.GET.getlist("name_category", [])
    supplier_names = request.GET.getlist("name_supplier", [])
    search_query = request.GET.get("searchInput")

    products = Products.objects.all().order_by("-id")
    user = request.user

    paginator = Paginator(products, 24)
    page_number = request.GET.get("page", 1)

    if category_names:
        category_filters = Q()
        for category_name in category_names:
            category_filters |= Q(categories__pk=category_name)
        products = products.filter(category_filters).order_by("-id")

    if supplier_names:
        supplier_filters = Q()
        for supplier_name in supplier_names:
            supplier_filters |= Q(suppliers__pk__in=supplier_name)
        products = products.filter(supplier_filters).distinct().order_by("-id")

    if search_query:
        products = products.filter(title__icontains=search_query).order_by("-id")

    if not (search_query or supplier_names or category_names):
        try:
            products = paginator.page(page_number)
        except PageNotAnInteger:
            products = paginator.page(1)
        except EmptyPage:
            products = paginator.page(paginator.num_pages)

    products_page_partial_html = render_to_string(
        "products/products_page_partial.html", {"products": products, "user": user}
    )
    return HttpResponse(products_page_partial_html)


# region Section CRUD: Create
@superuser_required
def create_product(request):
    if request.method == "POST":
        form = ProductsForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            uploaded_image = request.FILES.get("image")
            if uploaded_image:
                product.image = uploaded_image

            # Save the product first
            product.save()

            # Get the selected suppliers and their URLs and prices
            selected_suppliers = form.cleaned_data.get("suppliers")
            supplier_links = {}
            for supplier in selected_suppliers:
                supplier_link_key = f"supplierLink_{supplier.id}"
                supplier_price_key = f"supplierPrice_{supplier.id}"
                supplier_link_value = request.POST.get(supplier_link_key)
                supplier_price_value = request.POST.get(supplier_price_key)
                if supplier_link_value:
                    supplier_links[supplier] = (supplier_link_value, supplier_price_value)

            # Associate suppliers, URLs, and prices with the product
            for supplier, (link, price) in supplier_links.items():
                links_instance = Links.objects.create(url=link, supplier=supplier, price=price)
                product.links.add(links_instance)
                product.suppliers.add(supplier)

            messages.success(request, "Produto adicionado com sucesso!")
            return redirect("products_page")
        else:
            messages.error(request, "Erro ao criar produto. Verifique se as informações estão corretas.")
            return redirect("products_page")
    else:
        form = ProductsForm()

    return render(request, "products/products_page.html", {"form": form})


@superuser_required
def create_category(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.save()
            messages.success(request, "Categoria adicionada com sucesso!")
            return redirect("products_page")
        else:
            messages.error(request, "Erro ao criar Categoria. Verifique se as informações estão corretas.")
            return redirect("products_page")
    else:
        form = CategoryForm()

    return render(request, "products/products_page.html", {"form_categories": form})


@superuser_required
def create_supplier(request):
    if request.method == "POST":
        form = SupplierForm(request.POST, request.FILES)
        if form.is_valid():
            supplier = form.save(commit=False)
            uploaded_image = request.FILES.get("image")
            if uploaded_image:
                supplier.image_supplier = uploaded_image
            supplier.save()
            messages.success(request, "Fornecedor adicionado com sucesso!")
            return redirect("products_page")
        else:
            messages.error(request, "Erro ao criar Fornecedor. Verifique se as informações estão corretas.")
            return redirect("products_page")
    else:
        form = CategoryForm()

    return render(request, "products/products_page.html", {"form_supplier": form})


# endregion


# region Section CRUD: Update
@superuser_required
def update_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == "POST":
        updated_supplier_name = request.POST.get("update_supplier_name")
        supplier_image_name = "update_supplier_image" + str(supplier_id)
        updated_supplier_image = request.FILES.get(supplier_image_name)
        if updated_supplier_image:
            supplier.image_supplier = updated_supplier_image
        supplier.name_supplier = updated_supplier_name
        supplier.save()
        messages.success(request, "Fornecedor atualizado com sucesso!")
    else:
        messages.error(request, "Erro ao atualizar fornecedor.")
    return redirect("products_page")


@superuser_required
def update_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == "POST":
        updated_category_name = request.POST.get("update_category_name")
        category.name_category = updated_category_name
        category.save()
        messages.success(request, "Categoria atualizada com sucesso!")
    else:
        messages.error(request, "Erro ao atualizar categoria.")
    return redirect("products_page")


@superuser_required
def update_product(request, product_id):
    product = get_object_or_404(Products, id=product_id)
    if request.method == "POST":
        updated_product_title = request.POST.get("update_product_title")
        updated_product_image = request.FILES.get("update_product_image_" + str(product_id))
        product.title = updated_product_title
        if updated_product_image:
            product.image = updated_product_image
        category_id = request.POST.get("categories")
        category = Category.objects.get(id=category_id)
        product.categories = category
        product.save()
        supplier_ids = request.POST.getlist("suppliers")
        suppliers = Supplier.objects.filter(id__in=supplier_ids)
        product.suppliers.set(suppliers)
        product.links.clear()

        for supplier in suppliers:
            url = request.POST.get(f"supplierLink_{supplier.id}", "")
            price_str = request.POST.get(f"supplierPrice_{supplier.id}", "")
            if "," in price_str:
                price_str = price_str.replace(",", ".")
            try:
                price_decimal = Decimal(price_str)
            except (ValueError, DecimalException):
                pass
            link = Links.objects.create(url=url, supplier=supplier, price=price_decimal)
            product.links.add(link)

        messages.success(request, "Produto atualizado com sucesso!")
    else:
        messages.success(request, "Erro ao atualizar produto.")
    return redirect("products_page")


# endregion


# region Section CURD: Delete
@superuser_required
def delete_product(request, product_id):
    product = Products.objects.filter(id=product_id)
    if request.method == "POST":
        product.delete()
        return redirect("products_page")
    return redirect("products_page")


@superuser_required
def delete_category(request, category_id):
    category = Category.objects.get(id=category_id)
    products_count = Products.objects.filter(categories=category).count()
    if request.method == "POST":
        if products_count > 0:
            messages.error(
                request, "Enquanto exisiterem produtos com esta categoria associada, não será possível eliminá-la."
            )
        else:
            category.delete()
            return redirect("products_page")
    return redirect("products_page")


@superuser_required
def delete_supplier(request, supplier_id):
    supplier = Supplier.objects.get(id=supplier_id)
    products_count = Products.objects.filter(suppliers=supplier).count()
    if request.method == "POST":
        if products_count > 0:
            messages.error(
                request, "Enquanto existirem produtos associados a este fornecedor, não será possível apagá-lo."
            )
        else:
            supplier.delete()
            return redirect("products_page")
    return redirect("products_page")


# endregion


def _robots_for(host_url: str):
    """Fetch and parse robots.txt for the host serving the product images."""
    from scrappers import robots as robots_rules

    try:
        response = requests.get(
            urljoin(host_url, "/robots.txt"),
            headers={"User-Agent": settings.PRODUCT_IMPORT_USER_AGENT},
            timeout=15,
        )
        return robots_rules.parse(response.text if response.ok else "", settings.PRODUCT_IMPORT_USER_AGENT)
    except requests.RequestException:
        return robots_rules.parse("", settings.PRODUCT_IMPORT_USER_AGENT)


def local_images(category: str, count: int):
    """Reuse the images the scraper already downloaded, if they are on disk.

    scrappers/peixoto2.py writes <category>_products/image_N.jpg alongside the
    spreadsheet, numbered by row, at the delay the supplier asks for. Reading
    those makes the import instant and hits the network zero times; anything
    missing falls through to fetch_images below.
    """
    folder = Path(settings.PRODUCT_IMPORT_LOCAL_IMAGE_DIR) / f"{category}_products"
    found: list[tuple[bytes, str] | tuple[None, None]] = []
    for index in range(count):
        path = folder / f"image_{index + 1}.jpg"
        if path.is_file():
            found.append((path.read_bytes(), path.name))
        else:
            found.append((None, None))
    return found


def fetch_images(image_urls):
    """Download product images, obeying each host's robots.txt and Crawl-delay.

    The previous version fired every URL at once with asyncio.gather, ignoring
    the delay the source site asks for. Honouring it makes the work sequential,
    and at the 30s some sites request that is far longer than a request should
    block - so it runs against a time budget and reports what it could not
    fetch. Products are still created; they simply show the placeholder until
    their image is filled in.
    """
    results = [(None, None)] * len(image_urls)
    if not image_urls:
        return results, 0, 0

    session = requests.Session()
    session.headers["User-Agent"] = settings.PRODUCT_IMPORT_USER_AGENT
    rules_by_host: dict[str, object] = {}
    deadline = time.monotonic() + settings.PRODUCT_IMPORT_MAX_SECONDS
    last_request_at: dict[str, float] = {}
    skipped_budget = 0
    skipped_robots = 0

    for index, url in enumerate(image_urls):
        parsed = urlparse(url)
        host = f"{parsed.scheme}://{parsed.netloc}"
        if host not in rules_by_host:
            rules_by_host[host] = _robots_for(host)
        rules = rules_by_host[host]

        if not rules.allows(url):
            skipped_robots += 1
            continue

        delay = rules.crawl_delay if rules.crawl_delay is not None else settings.PRODUCT_IMPORT_MIN_DELAY
        delay = max(delay, settings.PRODUCT_IMPORT_MIN_DELAY)
        wait = delay - (time.monotonic() - last_request_at.get(host, 0.0))
        if time.monotonic() + max(wait, 0) > deadline:
            skipped_budget = len(image_urls) - index - skipped_robots
            break
        if wait > 0:
            time.sleep(wait)

        try:
            response = session.get(url, timeout=30)
            last_request_at[host] = time.monotonic()
            if response.status_code == 200:
                results[index] = (response.content, os.path.basename(parsed.path))
        except requests.RequestException as exc:
            print(f"Error downloading image {url}: {exc}")

    return results, skipped_robots, max(skipped_budget, 0)


@superuser_required
def upload_excel(request):
    try:
        product_names = []
        product_links = []
        product_prices = []
        product_images = []
        product_suppliers = []
        product_categories = []
        supplier_flag = False
        category_flag = False
        suppliers = Supplier.objects.all()
        categories = Category.objects.all()
        products = Products.objects.all()

        if request.method == "POST":
            uploaded_file = request.FILES.get("excel_file")

            if uploaded_file and uploaded_file.name.endswith(".xlsx"):
                wb = load_workbook(uploaded_file)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6, values_only=True):
                    for col_idx, cell_value in enumerate(row, start=1):
                        if cell_value:
                            if col_idx == 1:
                                product_names.append(cell_value)
                            elif col_idx == 2:
                                product_links.append(cell_value)
                            elif col_idx == 3:
                                normalized_value = cell_value.replace(",", ".")
                                price_numeric = re.search(r"\d+[\.,]?\d*", normalized_value)
                                if price_numeric:
                                    price_float = float(price_numeric.group().replace(",", "."))
                                    product_prices.append(price_float)
                            elif col_idx == 4:
                                product_images.append(cell_value)
                            elif col_idx == 5 and not category_flag:
                                category = categories.filter(name_category=cell_value).first()
                                if not category:
                                    categories.create(name_category=cell_value)
                                    category = categories.filter(name_category=cell_value).first()
                                product_categories.append(category)
                                category_flag = True
                                break
                            elif col_idx == 6 and not supplier_flag:
                                supplier = suppliers.filter(name_supplier=cell_value).first()
                                if not supplier:
                                    category.delete()
                                    messages.error(
                                        request,
                                        "Este fornecedor não existe. Crie antes o fornecedor "
                                        "para conseguir fazer upload de produtos.",
                                    )
                                    return redirect("products_page")
                                product_suppliers.append(supplier)
                                supplier_flag = True

                len_list_names = len(product_names) - 1
                product_suppliers.extend([supplier] * len_list_names)

                if len(product_names) == len(product_links) == len(product_prices) == len(product_images):
                    category_obj = product_categories[0] if product_categories else None

                    # Prefer the scraper's own downloads; only go to the network
                    # for rows it does not cover.
                    image_results = local_images(category_obj.name_category, len(product_images))
                    reused = sum(1 for content, _ in image_results if content)
                    missing = [
                        url if content is None else None for url, (content, _) in zip(product_images, image_results)
                    ]
                    if any(u for u in missing):
                        fetched, skipped_robots, skipped_budget = fetch_images([u or "" for u in missing])
                        for i, item in enumerate(fetched):
                            if item[0] is not None:
                                image_results[i] = item
                    else:
                        skipped_robots = skipped_budget = 0
                    supplier_obj = product_suppliers[0]

                    # One transaction: a failure part-way through no longer leaves
                    # the category's previous products deleted and nothing in their
                    # place.
                    with transaction.atomic():
                        products.filter(suppliers=supplier_obj, categories=category_obj).delete()

                        new_products = []
                        new_links = []
                        for name, link, price, (image_content, image_name) in zip(
                            product_names, product_links, product_prices, image_results
                        ):
                            stored = ""
                            if image_content:
                                stored = default_storage.save(
                                    f"products_images/{image_name}", ContentFile(image_content)
                                )
                            new_products.append(Products(title=name, categories=category_obj, image=stored))
                            new_links.append(Links(url=link, price=price, supplier=supplier_obj))

                        # bulk_create rather than a save per row; PostgreSQL returns
                        # the primary keys, so the join rows can be built in bulk too.
                        Products.objects.bulk_create(new_products)
                        Links.objects.bulk_create(new_links)

                        supplier_join = Products.suppliers.through
                        links_join = Products.links.through
                        supplier_join.objects.bulk_create(
                            [supplier_join(products_id=p.pk, supplier_id=supplier_obj.pk) for p in new_products]
                        )
                        links_join.objects.bulk_create(
                            [links_join(products_id=p.pk, links_id=lk.pk) for p, lk in zip(new_products, new_links)]
                        )

                    created = len(new_products)
                    detail = f"{created} produtos importados."
                    if reused:
                        detail += f" {reused} imagens reutilizadas do disco."
                    if skipped_robots:
                        detail += f" {skipped_robots} imagens ignoradas por robots.txt."
                    if skipped_budget:
                        detail += (
                            f" {skipped_budget} imagens não descarregadas dentro do limite de "
                            f"{settings.PRODUCT_IMPORT_MAX_SECONDS}s (Crawl-delay do fornecedor)."
                        )
                    messages.success(request, detail)
                else:
                    category.delete()
                    messages.error(
                        request,
                        "Certifique-se de que o documento tem o formato certo. "
                        "Existe pelo menos um produto em que falta informação.",
                    )
            else:
                messages.error(request, "Por favor, selecione um ficheiro Excel (.xlsx) para fazer o upload.")
        else:
            messages.error(
                request, "O método de requisição não é suportado. Use o método POST para fazer o upload do arquivo."
            )
    except Exception as e:
        messages.error(request, f"Ocorreu um erro ao processar o upload: {str(e)}")

    return redirect("products_page")
